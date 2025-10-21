# consort.py, Charlie Jordan, 10/7/2025
# sort and analyze contingency files

import os
import argparse
import re
import math
import configparser
import pandas as pd
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
import openpyxl.styles

from consort.contingency import Contingency
from consort.tools import dump_contingencies, get_cat_numbers

# process cmd line arg if it exists
parser = argparse.ArgumentParser()
parser.add_argument("ini_path", default='config.ini', nargs='?',
        help="The program's config file. See default (config.ini) for more details.")
args = parser.parse_args()
config_file_path = Path(args.ini_path)

# load configs from ini file
config_data = configparser.ConfigParser()
config_data.read(config_file_path)
BUS_FILE                 = Path(config_data["PATHS"]["BUS_FILE"])
OUTPUT_CON_PATH          = Path(config_data["PATHS"]["OUTPUT_CON_PATH"])
ALL_INPUT_CON_FILES_PATH = Path(config_data["PATHS"]["ALL_INPUT_CON_FILES_PATH"])
SHOW_INPUT_FILE_PROCESSING  = config_data["FLAGS"].getboolean("SHOW_INPUT_FILE_PROCESSING")
SHOW_LOADED_CON_SUMMARY     = config_data["FLAGS"].getboolean("SHOW_LOADED_CON_SUMMARY")
SHOW_DUP_ID_LIST            = config_data["FLAGS"].getboolean("SHOW_DUP_ID_LIST")
SHOW_NERC_CAT_SUMMARY       = config_data["FLAGS"].getboolean("SHOW_NERC_CAT_SUMMARY")
SHOW_DUP_LINE_COUNT_SUMMARY = config_data["FLAGS"].getboolean("SHOW_DUP_LINE_COUNT_SUMMARY")
SHOW_OUTPUT_FILE_PROGRESS   = config_data["FLAGS"].getboolean("SHOW_OUTPUT_FILE_PROGRESS")
WAIT_FOR_INPUT_TO_CLOSE     = config_data["FLAGS"].getboolean("WAIT_FOR_INPUT_TO_CLOSE")

# read buses from bus file
with open(BUS_FILE, 'r') as bus_file:
    BUSES = [x for x in bus_file.read().split('\n') if x]
    BUSES.sort()

# make sure output con directory exists
try:
    os.mkdir(OUTPUT_CON_PATH)
except FileExistsError:
    pass

simple_regex = re.compile(
    r'(?:\/\*.*\n){3}'
    r'\s*CONTINGENCY \'(.+)\'.*\n'
    r'([\n\S\s]*?)\s*END\n',
    re.MULTILINE
)

con_dict: dict[str, Contingency] = {}
con_set: set[Contingency] = set()
double_dict: dict[str, set[Contingency]] = {}
try:
    os.remove('filtered_con_text.con')
except OSError:
    pass
if SHOW_INPUT_FILE_PROCESSING:
    print("Input Files:")
for p in ALL_INPUT_CON_FILES_PATH.glob("**/*.con"):
    kbytes = os.stat(p).st_size/1024
    if SHOW_INPUT_FILE_PROCESSING:
        print(f'  {p.name} - ', end='')
    with open(p, 'r') as read_file, open('filtered_con_text.txt', 'a') as filter_file:
        text = read_file.read()

        # get all contingencies out of the files
        count = 0
        for m in re.finditer(simple_regex, text):
            contingency = Contingency(m[0], p.name)
            # store contingencies that have the same id but are otherwise
            # different. utilize set element uniqueness to avoid adding extras
            if contingency not in con_set and contingency.id in con_dict:
                if contingency.id not in double_dict:
                    double_dict[contingency.id] = set()
                double_dict[contingency.id].add(contingency)
                double_dict[contingency.id].add(con_dict[contingency.id])
            con_set.add(contingency)
            con_dict[contingency.id] = contingency
            count += 1
        if SHOW_INPUT_FILE_PROCESSING:
            print(count)

        # substitute out all contingencies and see what's left. should be nothing
        # but comments and some END keywords
        subbed_text = re.sub(simple_regex, '', text)
        # simplify huge comment blocks
        subbed_text = re.sub(r'(\/\* =+\n){2,}', '/* ' + ('='*105) + '+\n', subbed_text)
        filter_file.write(f'/* === FILE {p.name} ===\n')
        filter_file.write(subbed_text)

if SHOW_INPUT_FILE_PROCESSING:
    print()
if SHOW_LOADED_CON_SUMMARY:
    print("Loaded Contingencies Summary:")
    print(f"  {len(con_set):<6}", "Unique Contingencies (con_set)")
    print(f"  {len(con_dict):<6}", "Unique IDs (con_dict)")
    print(f"  {len(double_dict):<6}", "Contingencies with Duplicate ID's (double_dict)")

# separate doubled contingency id's
for id_, set_ in double_dict.items():
    sorted_cons = sorted(set_, key=lambda x: (x.lines_str, x.full_str))
    for i, con in enumerate(sorted_cons):
        con_set.remove(con)
        new_id = f"{id_}_{i}"
        con.change_id(new_id)
        con_set.add(con)

# create a set of contingencies only if they contain a bus in the supplied file
bus_filtered_con_set: set[Contingency] = set()
for i, x in enumerate(con_set):
    if any(f"BUS {bus}" in x.lines_str for bus in BUSES) and 'P6.' not in x.nerc_cat:
        bus_filtered_con_set.add(x)

if SHOW_LOADED_CON_SUMMARY:
    print(f"  {len(bus_filtered_con_set):<6}", "Unique Contingencies with Supplied Buses (bus_filtered_con_set)")
    print()

if SHOW_DUP_ID_LIST:
    print("Duplicate ID's:")
    if not double_dict:
        print("  NONE")
    for id_ in double_dict:
        print(f"  {id_}")
    print()

# dump all contingencies to a file
dump_contingencies(OUTPUT_CON_PATH / "All Filtered Contingencies.con", bus_filtered_con_set)

# show nerc category numerical breakdown
if SHOW_NERC_CAT_SUMMARY:
    print("NERC Category Summary:")
    cat_numbers = get_cat_numbers(bus_filtered_con_set)
    if not cat_numbers:
        print("  NONE")
    for cat, num in cat_numbers:
        print(f"  {num:<5} {cat}")
    print()

# get all the contingencies that share a set of statements
dup_lines_dict: dict[str, list[Contingency]] = {}
for x in bus_filtered_con_set:
    if x.lines_str not in dup_lines_dict:
        dup_lines_dict[x.lines_str] = []
    dup_lines_dict[x.lines_str].append(x)
dup_lines_dict = {line: ls for line, ls in dup_lines_dict.items() if len(ls) > 1}

# get the counts for how many duplicates there are and
# how many doubles, triples, etc.
if SHOW_DUP_LINE_COUNT_SUMMARY:
    dup_lines_count_dict = {}
    for line, ls in dup_lines_dict.items():
        len_ = len(ls)
        if len_ not in dup_lines_count_dict:
            dup_lines_count_dict[len_] = 0
        dup_lines_count_dict[len_] += 1

    print("Duplicate Lines Count Summary:")
    if not dup_lines_count_dict:
        print("  NONE")
    width = max(math.ceil(math.log10(x)) for x in dup_lines_count_dict.keys())
    for len_, count_ in sorted(dup_lines_count_dict.items(), key=lambda x: x[0]):
        print(f"  {len_:<{width}} - {count_}")
    print()

# collect the duplicated contingencies and store them in their sister
# contingencies' duplicates list
for i, (k, ls) in enumerate(dup_lines_dict.items()):
    for con in ls:
        other_cons = [c for c in ls if c != con]
        con.duplicates = other_cons

# group into files by (1.1), (1.2, 1.3, 2.1), (1.4), (7.1)
# (P2.2, P2.3, P2.4, P4, P5), (EE1, EE2, EE3)
# throw out 6.1s/6.*s (DB_ID_28234)
cats_to_search = [
    ('P1.1',),
    ('P1.2', 'P1.3', 'P2.1'),
    ('P1.4',),
    ('P7.1',),
    ('P2.2', 'P2.3', 'P2.4', 'P4.', 'P5.'),
    ('EE1', 'EE2', 'EE3')
]
output_file_names: list[str] = list(
    f"Group NERC {x}.con" for x in (
        'P1.1',
        'P1.2, 1.3, 2.1',
        'P1.4',
        'P7.1',
        'P2.2, 2.3, 2.4, 4, 5',
        'EE'
    )
)
output_file_groups = [[] for i in range(6)]
# group contingencies into which file they'll be in
for con in bus_filtered_con_set:
    for cat_list, group_list in zip(cats_to_search, output_file_groups):
        if any(cat in con.nerc_cat for cat in cat_list):
            group_list.append(con)
            break

# output the contingencies to their files

if SHOW_OUTPUT_FILE_PROGRESS:
    print("Output File Progress:")
for filename, group in zip(output_file_names, output_file_groups):
    if SHOW_OUTPUT_FILE_PROGRESS:
        print(f'  {filename}', end='', flush=True)
    dump_contingencies(OUTPUT_CON_PATH / filename, group)
    if SHOW_OUTPUT_FILE_PROGRESS:
        print(" - done")

# dump to excel file
# store contingency name and line data in contingency description in the lookup spreadsheet
XL_PATH = OUTPUT_CON_PATH / 'Contingency Lookup.xlsx'
if SHOW_OUTPUT_FILE_PROGRESS:
    print(f'  {XL_PATH.name}', end='', flush=True)
df = pd.DataFrame((x.make_csv_line_dict() for x in bus_filtered_con_set))
df.sort_values(['CONTINGENCY TYPE', 'CONTINGENCY'], axis=0, inplace=True)
df.to_excel(XL_PATH, index=False, sheet_name="Contingency Lookup")

wb = openpyxl.open(XL_PATH)
sheet = wb['Contingency Lookup']
for col in range(sheet.max_column):
    col_lens = []
    for row in range(sheet.max_row):
        cell = sheet.cell(row=row+1, column=col+1)
        cell.font = openpyxl.styles.Font(name='Consolas', size=9)
        cell.alignment = openpyxl.styles.Alignment(vertical='center', wrap_text=True)

        # measure cell size in column
        cell_val = str(cell.value)
        cell_len = max(len(x) for x in cell_val.split('\n'))
        col_lens.append(cell_len)
        # max_col_len = max(max_col_len, cell_len)
    max_col_len = max(sorted(col_lens)[:int(len(col_lens)*.95)])
    sheet.column_dimensions[get_column_letter(col+1)].width = max_col_len

# make the contingency lookup tab into a sortable table
table_ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
tab = Table(displayName="ContingencyLookup", ref=table_ref)
sheet.add_table(tab)

wb.save(XL_PATH)
if SHOW_OUTPUT_FILE_PROGRESS:
    print(' - done')
    print()

print("Done")
if WAIT_FOR_INPUT_TO_CLOSE:
    input("Press Enter to Exit")
