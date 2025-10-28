# tools.py, Charlie Jordan, 10/20/2025

import time
from typing import Callable, Iterable
from pathlib import Path
import pandas as pd

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
import openpyxl.styles

from .contingency import Contingency

SEP_LINE = '/* ' + '='*105 + '\n'

# dump all contingencies to a file
def dump_contingencies(file_name: str | Path, contingencies: Iterable[Contingency],
            sort_func: Callable[[Contingency],tuple] = lambda x: (x.submitter, x.nerc_cat, x.id)):
    with open(file_name, 'w') as file:
        # write file name and time at the top of the file
        file.write(f"/* {Path(file_name).name}, Generated {time.asctime()}\n")

        # write category summary at the top of the file
        file.write(SEP_LINE)
        cat_numbers = get_cat_numbers(contingencies)
        file.write('/* Category Count:\n')
        for cat, num in cat_numbers:
            file.write(f'/*\t\t{num:>4} - {cat}\n')

        # keep track of submitters to place contingency grouping comments
        prev_submitter = ''
        for con in sorted(contingencies, key=sort_func):
            if con.submitter != prev_submitter:
                prev_submitter = con.submitter
                file.write(SEP_LINE)
                file.write(
                    f'/* \t\t\t\t\tContingency Definitions Submitted by: {con.submitter}\n'
                )
            file.write(SEP_LINE)
            file.write(con.full_str)
        file.write(SEP_LINE)
        file.write("END\n")

def get_cat_numbers(ls: Iterable[Contingency]):
    cat_numbers = {}
    for con in ls:
        if con.nerc_cat not in cat_numbers:
            cat_numbers[con.nerc_cat] = 1
        else:
            cat_numbers[con.nerc_cat] += 1
    return tuple(sorted(cat_numbers.items()))

def dump_2_excel(output_path, con_set, show_file_progress):
    try:
        _dump_2_excel(output_path, con_set, show_file_progress)
    except PermissionError:
        if show_file_progress:
            print(' - FAILED!')
            print()
        print(f"ERROR: {output_path.name} is open in Excel, and is unable to be saved over.")
        print(f"  Please close the file and rerun this program.")
        print()
    else:
        if show_file_progress:
            print(' - done')
            print()

def _dump_2_excel(output_path, con_set, show_file_progress):
    if show_file_progress:
        print(f'  {output_path.name}', end='', flush=True)
    df = pd.DataFrame((x.make_csv_line_dict() for x in con_set))
    df.sort_values(['CONTINGENCY TYPE', 'CONTINGENCY'], axis=0, inplace=True)
    df.to_excel(output_path, index=False, sheet_name="Contingency Lookup")

    wb = openpyxl.open(output_path)
    sheet = wb['Contingency Lookup']
    for col in range(sheet.max_column):
        col_lens = []
        for row in range(sheet.max_row):
            cell = sheet.cell(row=row+1, column=col+1)
            cell.font = openpyxl.styles.Font(name='Consolas', size=9)
            cell.alignment = openpyxl.styles.Alignment(vertical='center', wrap_text=True)

            # measure cell size in column
            cell_val = str(cell.value)
            cell_len = max(len(x) for x in cell_val.split('\n'))
            col_lens.append(cell_len)
            # max_col_len = max(max_col_len, cell_len)
        max_col_len = max(sorted(col_lens)[:int(len(col_lens)*.95)])
        sheet.column_dimensions[get_column_letter(col+1)].width = max_col_len

    # make the contingency lookup tab into a sortable table
    table_ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    tab = Table(displayName="ContingencyLookup", ref=table_ref)
    sheet.add_table(tab)

    wb.save(output_path)
